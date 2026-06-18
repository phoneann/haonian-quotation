#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""好年企業有限公司 — 報價單產生介面（雲端版）"""

import os, io, tempfile
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pdf_generator import build_pdf, make_doc_number

st.set_page_config(
    page_title="好年報價單產生器",
    page_icon="📄",
    layout="centered",
    initial_sidebar_state="collapsed",
)
st.title("📄 好年企業 報價單產生器")

# ══════════════════════════════════════════
# 手機版 CSS 優化
# ══════════════════════════════════════════
st.markdown("""
<style>
/* 防止 iOS / Android 自動放大輸入框 */
input, textarea, select {
    font-size: 16px !important;
}

/* 所有按鈕：觸控友好最小高度 */
.stButton > button {
    min-height: 48px !important;
    font-size: 15px !important;
    touch-action: manipulation;
    -webkit-tap-highlight-color: transparent;
}

/* ── 手機（≤640px）── */
@media (max-width: 640px) {
    /* 縮減頁面邊距，讓內容更寬 */
    .main .block-container,
    [data-testid="stMainBlockContainer"] {
        padding-left: 0.6rem !important;
        padding-right: 0.6rem !important;
        padding-top: 0.5rem !important;
        max-width: 100% !important;
    }

    /* 所有 st.columns() 在手機上垂直堆疊 */
    [data-testid="stHorizontalBlock"] {
        flex-direction: column !important;
        gap: 0 !important;
    }
    [data-testid="stHorizontalBlock"] > div[data-testid="stColumn"],
    [data-testid="stHorizontalBlock"] > [data-testid="column"] {
        width: 100% !important;
        flex: none !important;
        min-width: 100% !important;
    }

    /* 標題字級縮小 */
    h1 { font-size: 1.35rem !important; line-height: 1.4 !important; }
    h2, h3 { font-size: 1.1rem !important; }

    /* 圖片不超出螢幕 */
    img { max-width: 100% !important; height: auto !important; }

    /* 刪除按鈕（🗑️ / ✕）不要因垂直堆疊顯得太寬 */
    [data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:last-child .stButton > button {
        min-height: 44px !important;
        width: 100% !important;
    }

    /* container border 內距 */
    [data-testid="stVerticalBlockBorderWrapper"] {
        padding: 0.5rem !important;
    }
}

/* ── 平板（641–768px）── */
@media (min-width: 641px) and (max-width: 768px) {
    .main .block-container,
    [data-testid="stMainBlockContainer"] {
        padding-left: 1rem !important;
        padding-right: 1rem !important;
    }
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════
# 工具函式
# ══════════════════════════════════════════
def export_excel(customer, groups):
    wb = Workbook()
    ws = wb.active
    ws.title = "報價資料"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    yellow = PatternFill("solid", fgColor="FFF2CC")
    blue   = PatternFill("solid", fgColor="DDEEFF")
    hdr    = PatternFill("solid", fgColor="C9DAF8")

    def lrow(r, label, value=""):
        ws.cell(r,1,label).font = Font(bold=True)
        ws.cell(r,1).fill = yellow
        ws.cell(r,2,value)

    ws.cell(1,1,"【客戶資訊】").font = Font(bold=True); ws.cell(1,1).fill = hdr
    lrow(2,"客戶名稱",customer["name"])
    lrow(3,"客戶編號",customer["number"])
    lrow(4,"統一編號",customer["tax_id"])
    lrow(5,"客戶電話",customer["phone"])
    lrow(6,"有效日期",customer["valid"])
    lrow(7,"報價日期",customer["date"])
    lrow(8,"序號",int(customer["seq"]))
    ws.cell(9,1,"【品項明細】").font = Font(bold=True); ws.cell(9,1).fill = hdr
    for c2,h in enumerate(["品名","數量","單價（元）","備註"],start=1):
        cell = ws.cell(10,c2,h)
        cell.font = Font(bold=True); cell.fill = blue
        cell.alignment = Alignment(horizontal="center")
    r = 11
    for g in groups:
        for row in g["rows"]:
            if g["name"].strip() or row["qty"].strip():
                ws.cell(r,1,g["name"])
                try: ws.cell(r,2, float(str(row["qty"]).replace(",","").strip()))
                except: ws.cell(r,2, str(row["qty"]))
                try: ws.cell(r,3, float(str(row["price"]).replace(",","").strip()))
                except: ws.cell(r,3, str(row["price"]))
                ws.cell(r,4,row["note"]); r+=1
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

def load_excel(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    def cell(r,c2):
        v = ws.cell(row=r,column=c2).value
        return str(v).strip() if v is not None else ""
    customer = {
        "name":cell(2,2),"number":cell(3,2),"tax_id":cell(4,2),
        "phone":cell(5,2),"valid":cell(6,2),"date":cell(7,2),
        "seq":int(cell(8,2)) if cell(8,2).isdigit() else 1,
    }
    raw=[]; r2=11
    while True:
        name=cell(r2,1)
        if not name: break
        raw.append({"name":name,"qty":cell(r2,2),"price":cell(r2,3),"note":cell(r2,4)}); r2+=1
    groups=[]
    for rr in raw:
        if groups and groups[-1]["name"]==rr["name"]:
            groups[-1]["rows"].append({"qty":rr["qty"],"price":rr["price"],"note":rr["note"]})
        else:
            groups.append({"name":rr["name"],"img_tmp":None,"rows":[{"qty":rr["qty"],"price":rr["price"],"note":rr["note"]}]})
    return customer, groups if groups else [blank_group()]

def blank_group(): return {"name":"","img_tmp":None,"rows":[{"qty":"","price":"","note":""}]}
def blank_row():   return {"qty":"","price":"","note":""}

# ══════════════════════════════════════════
# Session state
# ══════════════════════════════════════════
if "customer" not in st.session_state:
    st.session_state.customer = {"name":"","number":"","tax_id":"","phone":"","valid":"","date":"","seq":1}
if "groups" not in st.session_state:
    st.session_state.groups = [blank_group()]

# ══════════════════════════════════════════
# 側邊欄：匯入舊 Excel
# ══════════════════════════════════════════
with st.sidebar:
    st.header("📂 匯入舊報價單")
    excel_file = st.file_uploader("上傳 Excel 自動帶入", type=["xlsx"])
    if excel_file:
        try:
            c2, g2 = load_excel(excel_file.read())
            st.session_state.customer = c2
            st.session_state.groups   = g2
            st.success("已帶入！")
        except Exception as e:
            st.error(f"讀取失敗：{e}")

# ══════════════════════════════════════════
# ① 客戶資訊
# ══════════════════════════════════════════
st.subheader("① 客戶資訊")
c = st.session_state.customer
col1, col2 = st.columns(2)
with col1:
    c["name"]   = st.text_input("客戶名稱 *", c["name"])
    c["number"] = st.text_input("客戶編號", c["number"])
    c["valid"]  = st.text_input("有效日期（YYYY/MM/DD）", c["valid"])
    c["seq"]    = st.number_input("序號（當日第幾張）", min_value=1, max_value=99, value=int(c["seq"]))
with col2:
    c["phone"]  = st.text_input("客戶電話", c["phone"])
    c["tax_id"] = st.text_input("統一編號", c["tax_id"])
    c["date"]   = st.text_input("報價日期（YYYY/MM/DD）", c["date"])
try:
    st.caption(f"📋 單據編號：**{make_doc_number(c['date'], int(c['seq']))}**")
except: pass

st.markdown("---")

# ══════════════════════════════════════════
# ② 品項明細
# ══════════════════════════════════════════
st.subheader("② 品項明細")
groups = st.session_state.groups
del_group = None
del_row   = None

for gi, g in enumerate(groups):
    with st.container(border=True):
        h1, h2 = st.columns([5,1])
        g["name"] = h1.text_input("品名", g["name"], key=f"gname_{gi}", placeholder="例：繩索鐵環")
        if gi > 0:
            if h2.button("🗑️", key=f"delg_{gi}", help="刪除此產品"):
                del_group = gi

        img_up = st.file_uploader("產品圖片（截圖直接拖進來）", type=["png","jpg","jpeg"], key=f"img_{gi}")
        if img_up:
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp.write(img_up.read()); tmp.close()
            g["img_tmp"] = tmp.name
        if g.get("img_tmp") and os.path.exists(g["img_tmp"]):
            st.image(g["img_tmp"], width=180)

        st.caption("數量（數字或文字如「1批」） ／ 單價 ／ 備註")
        for ri, row in enumerate(g["rows"]):
            rc = st.columns([2,2,4,1])
            row["qty"]   = rc[0].text_input("數量",  row["qty"],  key=f"qty_{gi}_{ri}",   label_visibility="collapsed", placeholder="10000 或 1批")
            row["price"] = rc[1].text_input("單價",  row["price"],key=f"price_{gi}_{ri}",  label_visibility="collapsed")
            row["note"]  = rc[2].text_input("備註",  row["note"], key=f"note_{gi}_{ri}",   label_visibility="collapsed")
            if ri > 0:
                if rc[3].button("✕", key=f"delr_{gi}_{ri}"): del_row=(gi,ri)
            else: rc[3].write("")

        if st.button("＋ 加一組數量", key=f"addrow_{gi}"):
            g["rows"].append(blank_row()); st.rerun()

if del_group is not None:
    st.session_state.groups.pop(del_group); st.rerun()
if del_row is not None:
    gi2,ri2=del_row; st.session_state.groups[gi2]["rows"].pop(ri2); st.rerun()

if st.button("＋ 新增產品"):
    st.session_state.groups.append(blank_group()); st.rerun()

st.markdown("---")

# ══════════════════════════════════════════
# ③ 交易條件備註
# ══════════════════════════════════════════
st.subheader("③ 交易條件備註")
if "remarks" not in st.session_state:
    st.session_state.remarks = ""
st.session_state.remarks = st.text_area(
    "填寫交期、付款條件等（會顯示在報價單底部）",
    st.session_state.remarks,
    placeholder="例：交期：接單後一個月內出貨。付款條件：月結30天。",
    height=100,
)

st.markdown("---")

# ══════════════════════════════════════════
# ④ 產出
# ══════════════════════════════════════════
st.subheader("④ 產出")
customer_data = {k: str(v) for k,v in c.items()}
date_tag = c["date"].replace("/","").replace("-","")[2:]
out_name = f"{c['name']}_報價單_{date_tag}"

all_items = []
for g in groups:
    for row in g["rows"]:
        if g["name"].strip() or row["qty"].strip():
            all_items.append({"name":g["name"],"qty":row["qty"],"price":row["price"],
                               "note":row["note"],"_img":g.get("img_tmp") or ""})

col_pdf, col_xl = st.columns(2)

with col_pdf:
    if st.button("🖨️ 產生 PDF", type="primary", use_container_width=True):
        if not c["name"]:
            st.error("請填客戶名稱")
        elif not all_items:
            st.error("請填至少一筆品項")
        else:
            tmp_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
            tmp_pdf.close()
            try:
                build_pdf({"customer":customer_data,"image_path":"","items":all_items,"remarks":st.session_state.remarks}, tmp_pdf.name)
                pdf_bytes = open(tmp_pdf.name,"rb").read()
                os.unlink(tmp_pdf.name)
                st.session_state["pdf_bytes"] = pdf_bytes
                st.session_state["pdf_name"]  = out_name+".pdf"
                st.success("✅ PDF 產生成功！")
            except Exception as e:
                st.error(f"失敗：{e}")
    if "pdf_bytes" in st.session_state:
        st.download_button("⬇️ 下載 PDF", st.session_state["pdf_bytes"],
                           file_name=st.session_state["pdf_name"],
                           mime="application/pdf", use_container_width=True)

with col_xl:
    if st.button("📊 存成 Excel", use_container_width=True):
        if not c["name"]:
            st.error("請填客戶名稱")
        else:
            xl_bytes = export_excel(customer_data, groups)
            st.session_state["xl_bytes"] = xl_bytes
            st.session_state["xl_name"]  = out_name+".xlsx"
            st.success("✅ Excel 已準備好！")
    if "xl_bytes" in st.session_state:
        st.download_button("⬇️ 下載 Excel", st.session_state["xl_bytes"],
                           file_name=st.session_state["xl_name"],
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)
